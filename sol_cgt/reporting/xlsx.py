"""XLSX export helpers."""
from __future__ import annotations

from decimal import Decimal
import logging
from pathlib import Path
from typing import Iterable, Sequence
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .. import utils
from ..pricing import AudPriceProvider
from ..types import AcquisitionLot, DisposalRecord, LotMoveRecord, NormalizedEvent, WarningRecord
from .schema import SUMMARY_BY_TOKEN_COLUMNS, WALLET_SUMMARY_COLUMNS


def _apply_header_style(sheet) -> None:
    bold = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"


def _auto_width(sheet) -> None:
    for col in sheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in col]
        max_len = max((len(value) for value in values), default=0)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _format_numbers(sheet, columns: Iterable[str], fmt: str) -> None:
    header = [cell.value for cell in sheet[1]]
    col_idx = {name: idx + 1 for idx, name in enumerate(header)}
    for name in columns:
        idx = col_idx.get(name)
        if not idx:
            continue
        for row in sheet.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                cell.number_format = fmt


def _transaction_rows(events: Sequence[NormalizedEvent], price_provider: AudPriceProvider) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    logger = logging.getLogger(__name__)
    for event in sorted(events, key=lambda ev: (ev.ts, ev.raw.get("signature") or ev.id, ev.id)):
        mint = None
        symbol = None
        qty = None
        value_aud = None
        price_aud = None
        if event.base_token is not None:
            token = event.base_token
            mint = token.mint
            symbol = token.symbol
            qty = token.amount
            if "proceeds_aud" in event.raw:
                value_aud = Decimal(str(event.raw["proceeds_aud"]))
            elif "proceeds_hint_aud" in event.raw:
                value_aud = Decimal(str(event.raw["proceeds_hint_aud"]))
            else:
                price = price_provider.price_aud(token.mint, event.ts, context=event.raw)
                if price is None:
                    logger.warning("Missing price for mint=%s at %s", token.mint, event.ts.isoformat())
                    value_aud = Decimal("0")
                else:
                    value_aud = utils.quantize_aud(price * token.amount)
            price_aud = utils.quantize_aud(value_aud / token.amount) if token.amount else Decimal("0")
        elif event.quote_token is not None:
            token = event.quote_token
            mint = token.mint
            symbol = token.symbol
            qty = token.amount
            if "cost_aud" in event.raw:
                value_aud = Decimal(str(event.raw["cost_aud"]))
            elif "cost_hint_aud" in event.raw:
                value_aud = Decimal(str(event.raw["cost_hint_aud"]))
            else:
                price = price_provider.price_aud(token.mint, event.ts, context=event.raw)
                if price is None:
                    logger.warning("Missing price for mint=%s at %s", token.mint, event.ts.isoformat())
                    value_aud = Decimal("0")
                else:
                    value_aud = utils.quantize_aud(price * token.amount)
            price_aud = utils.quantize_aud(value_aud / token.amount) if token.amount else Decimal("0")
        fee_aud = Decimal(str(event.raw.get("fee_aud", "0")))
        rows.append(
            {
                "date_time_local": utils.to_au_local(event.ts).isoformat(),
                "signature": event.raw.get("signature") or event.id.split("#")[0],
                "wallet": event.wallet,
                "event_type": event.kind,
                "mint": mint or "",
                "symbol": symbol or "",
                "quantity": str(qty or Decimal("0")),
                "price_aud": str(price_aud or Decimal("0")),
                "value_aud": str(value_aud or Decimal("0")),
                "fee_aud": str(fee_aud),
                "counterparty": event.counterparty or "",
                "notes": ",".join(sorted(event.tags)) if event.tags else "",
                "source": event.raw.get("source") or "",
            }
        )
    return rows


def _lot_rows(lots: Sequence[AcquisitionLot]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for lot in lots:
        rows.append(
            {
                "lot_id": lot.lot_id,
                "wallet": lot.wallet,
                "mint": lot.token_mint,
                "symbol": lot.token_symbol or "",
                "acquisition_date_local": utils.to_au_local(lot.ts).isoformat(),
                "qty_acquired": str(lot.qty_acquired),
                "unit_cost_aud": str(lot.unit_cost_aud),
                "remaining_qty": str(lot.remaining_qty),
                "source_event": lot.source_event or "",
            }
        )
    return rows


def _disposal_rows(disposals: Sequence[DisposalRecord]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record in disposals:
        rows.append(
            {
                "disposal_date_local": utils.to_au_local(record.ts).isoformat(),
                "wallet": record.wallet,
                "mint": record.token_mint,
                "symbol": record.token_symbol or "",
                "qty": str(record.qty_disposed),
                "proceeds_aud": str(record.proceeds_aud),
                "cost_base_aud": str(record.cost_base_aud),
                "gain_loss_aud": str(record.gain_loss_aud),
                "held_days": str(record.held_days),
                "discount_eligible": str(record.long_term),
                "method": record.method,
                "signature": record.signature or record.event_id.split("#")[0],
                "notes": record.notes or "",
            }
        )
    return rows


def _lot_move_rows(lot_moves: Sequence[LotMoveRecord]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for move in lot_moves:
        rows.append(
            {
                "tx_signature": move.tx_signature,
                "date_time_local": utils.to_au_local(move.ts).isoformat(),
                "src_wallet": move.src_wallet,
                "dst_wallet": move.dst_wallet,
                "mint": move.mint,
                "symbol": move.symbol or "",
                "amount": str(move.amount),
                "fee_aud": str(move.fee_aud),
                "lots_consumed": utils.json_dumps(move.lots_consumed),
                "lots_created": utils.json_dumps(move.lots_created),
            }
        )
    return rows


def export_xlsx(
    path: Path,
    *,
    overview: dict[str, str],
    events: Sequence[NormalizedEvent],
    lots: Sequence[AcquisitionLot],
    disposals: Sequence[DisposalRecord],
    summary_by_token: Sequence[dict[str, object]],
    wallet_summary: Sequence[dict[str, object]],
    lot_moves: Sequence[LotMoveRecord],
    warnings: Sequence[WarningRecord],
    price_provider: AudPriceProvider,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    overview_sheet.append(["Metric", "Value"])
    for key, value in overview.items():
        overview_sheet.append([key, value])
    _apply_header_style(overview_sheet)
    _auto_width(overview_sheet)

    tx_sheet = workbook.create_sheet("Transactions")
    tx_rows = _transaction_rows(events, price_provider)
    if tx_rows:
        tx_sheet.append(list(tx_rows[0].keys()))
        for row in tx_rows:
            tx_sheet.append(list(row.values()))
        _apply_header_style(tx_sheet)
        _format_numbers(tx_sheet, ["price_aud", "value_aud", "fee_aud"], "#,##0.00")
        _format_numbers(tx_sheet, ["quantity"], "#,##0.000000000")
        _auto_width(tx_sheet)

    lots_sheet = workbook.create_sheet("Lots")
    lot_rows = _lot_rows(lots)
    if lot_rows:
        lots_sheet.append(list(lot_rows[0].keys()))
        for row in lot_rows:
            lots_sheet.append(list(row.values()))
        _apply_header_style(lots_sheet)
        _format_numbers(lots_sheet, ["unit_cost_aud"], "#,##0.00")
        _format_numbers(lots_sheet, ["remaining_qty", "qty_acquired"], "#,##0.000000000")
        _auto_width(lots_sheet)

    disposals_sheet = workbook.create_sheet("Disposals")
    disposal_rows = _disposal_rows(disposals)
    if disposal_rows:
        disposals_sheet.append(list(disposal_rows[0].keys()))
        for row in disposal_rows:
            disposals_sheet.append(list(row.values()))
        _apply_header_style(disposals_sheet)
        _format_numbers(disposals_sheet, ["proceeds_aud", "cost_base_aud", "gain_loss_aud"], "#,##0.00")
        _format_numbers(disposals_sheet, ["qty"], "#,##0.000000000")
        _auto_width(disposals_sheet)

    summary_token_sheet = workbook.create_sheet("Summary by token")
    if summary_by_token:
        summary_token_sheet.append(SUMMARY_BY_TOKEN_COLUMNS)
        for row in summary_by_token:
            summary_token_sheet.append([row.get(col) for col in SUMMARY_BY_TOKEN_COLUMNS])
        _apply_header_style(summary_token_sheet)
        _auto_width(summary_token_sheet)

    wallet_sheet = workbook.create_sheet("Wallet summary")
    if wallet_summary:
        wallet_sheet.append(WALLET_SUMMARY_COLUMNS)
        for row in wallet_summary:
            wallet_sheet.append([row.get(col) for col in WALLET_SUMMARY_COLUMNS])
        _apply_header_style(wallet_sheet)
        _auto_width(wallet_sheet)

    lot_moves_sheet = workbook.create_sheet("Lot moves")
    move_rows = _lot_move_rows(lot_moves)
    if move_rows:
        lot_moves_sheet.append(list(move_rows[0].keys()))
        for row in move_rows:
            lot_moves_sheet.append(list(row.values()))
        _apply_header_style(lot_moves_sheet)
        _auto_width(lot_moves_sheet)

    warnings_sheet = workbook.create_sheet("Warnings")
    if warnings:
        warnings_sheet.append(list(warnings[0].model_dump().keys()))
        for warning in warnings:
            row = []
            for value in warning.model_dump().values():
                if hasattr(value, "isoformat"):
                    row.append(value.isoformat())
                else:
                    row.append(value)
            warnings_sheet.append(row)
        _apply_header_style(warnings_sheet)
        _auto_width(warnings_sheet)

    workbook.save(path)
