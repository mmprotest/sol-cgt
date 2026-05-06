from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook

from sol_cgt.accounting.engine import AccountingEngine, SimplePriceProvider
from sol_cgt.reporting import xlsx
from sol_cgt.types import MissingLotIssue, NormalizedEvent, TokenAmount


class DummyPriceProvider:
    def price_aud(self, mint: str, ts: datetime, *, context: dict | None = None) -> Decimal:
        return Decimal("1")


def test_missing_lots_xlsx_sheet(tmp_path) -> None:
    ts = datetime(2024, 1, 2, tzinfo=timezone.utc)
    token = TokenAmount(mint="TOKEN", amount_raw=5, decimals=0, symbol="TKN")
    event = NormalizedEvent(
        id="tx1#0",
        ts=ts,
        kind="sell",
        base_token=token,
        quote_token=None,
        fee_sol=Decimal("0"),
        wallet="W1",
        counterparty=None,
        raw={"signature": "sig1"},
        tags=set(),
    )
    engine = AccountingEngine(price_provider=SimplePriceProvider({"TOKEN": Decimal("1")}))
    missing_issues: list[MissingLotIssue] = []
    result = engine.process([event], strict_lots=False, missing_lot_issues=missing_issues)
    assert len(missing_issues) == 1
    assert result.acquisitions == []
    assert result.disposals == []

    xlsx_path = tmp_path / "report.xlsx"
    xlsx.export_xlsx(
        xlsx_path,
        overview={"Financial year": "all"},
        events=[event],
        lots=result.acquisitions,
        disposals=result.disposals,
        summary_by_token=[],
        wallet_summary=[],
        lot_moves=[],
        warnings=result.warnings,
        missing_lots=missing_issues,
        price_provider=DummyPriceProvider(),
    )

    workbook = load_workbook(xlsx_path)
    assert "Missing lots" in workbook.sheetnames
    sheet = workbook["Missing lots"]
    assert "missing history" in str(sheet["A1"].value).lower()
    assert sheet["A2"].value == "wallet"


def test_missing_lot_does_not_stop_later_events() -> None:
    ts = datetime(2024, 1, 2, tzinfo=timezone.utc)
    sell_missing = NormalizedEvent(
        id="tx1#0",
        ts=ts,
        kind="sell",
        base_token=TokenAmount(mint="TOKEN", amount_raw=5, decimals=0, symbol="TKN"),
        fee_sol=Decimal("0"),
        wallet="W1",
        raw={"signature": "sig1", "proceeds_aud": "5"},
    )
    buy = NormalizedEvent(
        id="tx2#0",
        ts=ts.replace(hour=1),
        kind="buy",
        quote_token=TokenAmount(mint="TOKEN2", amount_raw=10, decimals=0, symbol="TK2"),
        fee_sol=Decimal("0"),
        wallet="W1",
        raw={"signature": "sig2", "cost_aud": "10"},
    )
    engine = AccountingEngine(price_provider=SimplePriceProvider({}))
    issues: list[MissingLotIssue] = []
    result = engine.process([sell_missing, buy], strict_lots=False, missing_lot_issues=issues)
    assert len(issues) == 1
    assert any(lot.source_event == "tx2#0" for lot in result.acquisitions)
