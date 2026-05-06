from __future__ import annotations

import builtins
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook
import pytest

from sol_cgt.reporting import formats, summaries, xlsx
from sol_cgt.reporting.schema import (
    ACQUISITION_COLUMNS,
    DISPOSAL_COLUMNS,
    SUMMARY_BY_TOKEN_COLUMNS,
    SUMMARY_OVERALL_COLUMNS,
    WALLET_SUMMARY_COLUMNS,
)
from sol_cgt.types import AcquisitionLot, DisposalRecord, NormalizedEvent, TokenAmount, WarningRecord


class DummyPriceProvider:
    def price_aud(self, mint: str, ts: datetime, *, context: dict | None = None) -> Decimal:
        return Decimal("1")


def _sample_records() -> tuple[
    list[NormalizedEvent],
    list[AcquisitionLot],
    list[DisposalRecord],
    list[WarningRecord],
]:
    ts = datetime(2024, 1, 1, tzinfo=timezone.utc)
    lot = AcquisitionLot(
        lot_id="L1",
        wallet="W1",
        ts=ts,
        token_mint="M1",
        token_symbol="TOK",
        qty_acquired=Decimal("1"),
        unit_cost_aud=Decimal("10"),
        fees_aud=Decimal("0.10"),
        remaining_qty=Decimal("1"),
        source_event="sig1",
        source_type="swap",
    )
    disposal = DisposalRecord(
        event_id="E1",
        wallet="W1",
        ts=ts,
        token_mint="M1",
        token_symbol="TOK",
        qty_disposed=Decimal("1"),
        proceeds_aud=Decimal("12"),
        cost_base_aud=Decimal("10"),
        fees_aud=Decimal("0.10"),
        gain_loss_aud=Decimal("1.90"),
        long_term=False,
        held_days=10,
        method="FIFO",
        signature="sig1",
        notes="note",
    )
    event = NormalizedEvent(
        id="sig1#0",
        ts=ts,
        kind="sell",
        base_token=TokenAmount(
            mint="M1",
            symbol="TOK",
            decimals=6,
            amount_raw=1_000_000,
            amount=Decimal("1"),
        ),
        wallet="W1",
        raw={"signature": "sig1", "proceeds_aud": "12.00"},
    )
    warning = WarningRecord(
        ts=ts,
        wallet="W1",
        signature="sig1",
        code="warn",
        message="test warning",
    )
    return [event], [lot], [disposal], [warning]


def test_csv_and_xlsx_outputs(tmp_path) -> None:
    events, acquisitions, disposals, warnings = _sample_records()
    summary_by_token = summaries.summarize_by_token(disposals)
    summary_overall = summaries.summarize_overall(disposals)
    wallet_summary = summaries.summarize_by_wallet(disposals)

    formats.export_reports(
        tmp_path,
        acquisitions,
        disposals,
        summary_by_token,
        summary_overall,
        fmt="csv",
    )

    assert (tmp_path / "acquisitions.csv").read_text(encoding="utf-8").splitlines()[0] == ",".join(
        ACQUISITION_COLUMNS
    )
    assert (tmp_path / "disposals.csv").read_text(encoding="utf-8").splitlines()[0] == ",".join(
        DISPOSAL_COLUMNS
    )
    assert (tmp_path / "summary_by_token.csv").read_text(encoding="utf-8").splitlines()[0] == ",".join(
        SUMMARY_BY_TOKEN_COLUMNS
    )
    assert (tmp_path / "summary_overall.csv").read_text(encoding="utf-8").splitlines()[0] == ",".join(
        SUMMARY_OVERALL_COLUMNS
    )

    xlsx_path = tmp_path / "report.xlsx"
    xlsx.export_xlsx(
        xlsx_path,
        overview={"Financial year": "2024-2025"},
        events=events,
        lots=acquisitions,
        disposals=disposals,
        summary_by_token=summary_by_token,
        wallet_summary=wallet_summary,
        lot_moves=[],
        warnings=warnings,
        missing_lots=[],
        price_provider=DummyPriceProvider(),
    )

    workbook = load_workbook(xlsx_path)
    assert "Overview" in workbook.sheetnames
    assert "Missing lots" in workbook.sheetnames
    assert "Summary by token" in workbook.sheetnames
    assert "internal_transfers" in workbook.sheetnames
    assert "Wallet summary" in workbook.sheetnames
    assert [cell.value for cell in workbook["Summary by token"][1]] == SUMMARY_BY_TOKEN_COLUMNS
    assert [cell.value for cell in workbook["Wallet summary"][1]] == WALLET_SUMMARY_COLUMNS
    tx_headers = [cell.value for cell in workbook["Transactions"][1]]
    assert "valuation_method" in tx_headers
    assert "valuation_reference_asset" in tx_headers
    lots_headers = [cell.value for cell in workbook["Lots"][1]]
    unit_cost_col = lots_headers.index("unit_cost_aud") + 1
    assert workbook["Lots"].cell(row=2, column=unit_cost_col).number_format == "#,##0.000000000000"


def test_parquet_requires_extra(monkeypatch, tmp_path) -> None:
    original_import = builtins.__import__

    def _blocked_import(name: str, *args, **kwargs):
        if name.startswith("pyarrow"):
            raise ImportError("pyarrow missing")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", _blocked_import)

    with pytest.raises(RuntimeError, match="Parquet support requires the 'parquet' extra"):
        formats.write_parquet(tmp_path / "summary.parquet", [], columns=SUMMARY_OVERALL_COLUMNS)


def test_transaction_summary_groups_by_signature() -> None:
    ts = datetime(2024, 1, 1, tzinfo=timezone.utc)
    ev1 = NormalizedEvent(id="sigx#0", ts=ts, kind="swap", wallet="W1", raw={"signature": "sigx"})
    ev2 = NormalizedEvent(id="sigx#1", ts=ts, kind="transfer_out", wallet="W1", raw={"signature": "sigx"})
    rows = summaries.build_transaction_summary([ev1, ev2])
    assert len(rows) == 1
    assert rows[0]["signature"] == "sigx"
    assert rows[0]["classification"] == "trade"


def test_excel_safe_handles_nested_decimal() -> None:
    value = {"nested": [Decimal("1.25")]}
    safe = xlsx._excel_safe(value)
    assert isinstance(safe, str)
    assert "1.25" in safe


def test_export_xlsx_normalized_events_debug_nested_decimal_and_missing_price_aggregation(tmp_path, caplog) -> None:
    class MissingPriceProvider:
        def price_aud(self, mint: str, ts: datetime, *, context: dict | None = None):
            return None

    ts = datetime(2024, 1, 1, tzinfo=timezone.utc)
    event = NormalizedEvent(
        id="sig2#0",
        ts=ts,
        kind="sell",
        base_token=TokenAmount(mint="MISSING1", symbol="TOK", decimals=6, amount_raw=1_000_000, amount=Decimal("1")),
        wallet="W1",
        raw={"signature": "sig2"},
    )
    xlsx_path = tmp_path / "report_nested.xlsx"
    with caplog.at_level("WARNING"):
        xlsx.export_xlsx(
            xlsx_path,
            overview={"Financial year": "2024-2025"},
            events=[event],
            lots=[],
            disposals=[],
            summary_by_token=[],
            wallet_summary=[],
            lot_moves=[],
            warnings=[],
            missing_lots=[],
            price_provider=MissingPriceProvider(),
            normalized_events_debug=[{"payload": {"amount": Decimal("2.5")}}],
        )
    assert xlsx_path.exists()
    missing_logs = [r.message for r in caplog.records if "Missing prices in XLSX export" in r.message]
    assert len(missing_logs) == 1
    per_event_logs = [r.message for r in caplog.records if "Missing price for mint=" in r.message]
    assert per_event_logs == []
